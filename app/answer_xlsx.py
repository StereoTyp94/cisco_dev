import openpyxl, os
from app import db
from app.models import Product, Service
from datetime import date

header = ['Part#', 'SKU', 'Price $', 'Commentary']
services = {
    'snt':['SNT', 'ECDN', 'ECMU'],
    'snte':['SNTE', 'ECEN', 'ECMU'],
    'sntp':['SNTP', 'EC4N', 'ECMU']
            }


def query_con(sku, serv_lev):
    result = []
    if 'CON-' in sku:
        result = db.session.query(Product.part).join(Service).\
            filter(Service.sku == sku).first()
    else:
        result.append(sku)
    if result:
        return query_serv(result[0], serv_lev)
    return None


def query_serv(part, serv_lev):
    if serv_lev in services:
        for serv in services[serv_lev]:
            result = db.session.query(Product.part, Service.sku, Service.serv_gpl).join(Service). \
                filter(Service.serv_lev == serv, Product.part == part).first()
            if result:
                return result
    return None


def create_answer_excel(part_str, serv_lev):
    # if 'answer.xlsx' in os.listdir(os.path.join(os.getcwd(), 'excel_files')):
    #     os.remove(os.path.join(os.getcwd(), 'excel_files', 'answer.xlsx')) #удаляем файл
    parts = part_str.split()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    line = 2
    for part in parts:
        result = query_con(part, serv_lev)
        if result:
            ws.cell(row=line, column=1).value = result[0]
            ws.cell(row=line, column=2).value = result[1]
            ws.cell(row=line, column=3).value = result[2]
        else:
            ws.cell(row=line, column=1).value = part
            ws.cell(row=line, column=4).value = 'Оборудование EndOfSupport или не имеет отдельного смартнета.'
        line += 1
    filename = str(date.today()) + '_smartnet.xlsx'
    wb.save(os.path.join(os.getcwd(), 'excel_files', filename))
    return filename


def create_answer_table(part_str, serv_lev):
    table_answer = []
    table_answer.append(header)
    parts = part_str.split()
    for part in parts:
        result = query_con(part, serv_lev)
        if result:
            table_answer.append([result[0], result[1], result[2], ''])
        else:
            table_answer.append([part, 'N\A', 'N\A', 'Оборудование EndOfSupport или не имеет отдельного смартнета.'])

    return table_answer


# db.session.query(Product.part, Service.sku, Service.serv_gpl).join(Service)\
#     .filter(Service.serv_lev == 'SNT', Product.part == '01FT600X').all()
# db.session.query(Product.part, Service.sku, Service.serv_gpl).join(Service)\
#     .filter(Service.sku == 'CON-SNT-WSC2FPDL').first()