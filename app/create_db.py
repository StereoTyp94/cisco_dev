import openpyxl, os
from app import app, db
from app.models import Product, Service


def getParamExsel(sheet):
    for part_y in range(1, 20): #ищем начало таблицы и номера столбцов
        if 'Major' in str(sheet.cell(row=part_y, column=1).value):
            for part_x in range(1,15):
                if 'Product SKU' == str(sheet.cell(row=part_y, column=part_x).value):
                    break
            for serv_lev in range(1,15):
                if 'Service Level' == str(sheet.cell(row=part_y, column=serv_lev).value):
                    break
            for sku in range(1,15):
                if 'Service SKU' == str(sheet.cell(row=part_y, column=sku).value):
                    gpl = sku + 1
                    serv_gpl = gpl + 1
                    break
            break
    return part_y, part_x, serv_lev, sku, gpl, serv_gpl


def cleanDB():
    db.session.query(Product).delete()
    db.session.query(Service).delete()
    return True


def insertDB(sheet):
    part = ''
    count = 0
    print('Начал процедуру заполнения базы данных')
    for line in range(part_y + 2, end_table):
        if line * 100 // end_table > count:
            count = line * 100 // end_table
            print(str(count) + '%')
        if str(sheet.cell(row=line, column=part_x).value) != part:
            part = str(sheet.cell(row=line, column=part_x).value)
            if Product.query.filter_by(part=part).first() is None:
                try:
                    gpl_product = float(sheet.cell(row=line, column=gpl).value)
                except Exception:
                    gpl_product = -1
                product = Product(part=part, gpl=gpl_product)
                db.session.add(product)
                #db.session.commit()
            else:
                part_cur = Product.query.filter_by(part=part).first()
                part_cur.gpl = float(sheet.cell(row=line, column=gpl).value)
                #db.session.commit()
        sku_name = str(sheet.cell(row=line, column=sku).value)
        service = Service(serv_lev=sheet.cell(row=line, column=serv_lev).value, sku=sku_name, \
                          serv_gpl=float(sheet.cell(row=line, column=serv_gpl).value),
                          equipment=Product.query.filter_by(part=part).first())
        db.session.add(service)


        # if Service.query.filter_by(sku=sku_name).first() is None:
        #     service = Service(serv_lev=sheet.cell(row=line, column=serv_lev).value, sku=sku_name, \
        #                       serv_gpl=float(sheet.cell(row=line, column=serv_gpl).value),
        #                       equipment=Product.query.filter_by(part=part).first())
        #     db.session.add(service)
        #     #db.session.commit()
        # else:
        #     sku_cur = Service.query.filter_by(sku=sku_name).first()
        #     sku_cur.serv_gpl = float(sheet.cell(row=line, column=serv_gpl).value)
        #     #db.session.commit()
    db.session.commit()
    return True

print(os.getcwd())
path = os.chdir('..')
print(os.listdir())
print(os.path.join(os.getcwd(), 'price', 'Report1.xlsx'))
print(os.getcwd() + '\price\Report1.xlsx')


wb = openpyxl.load_workbook(os.path.join(os.getcwd(), 'price', 'Report.xlsx'))
print(True)
sheet = wb.get_active_sheet()
part_y, part_x, serv_lev, sku, gpl, serv_gpl = getParamExsel(sheet)
if part_y == 20:
    print('Не смог найти начало таблицы')
    exit()
end_table = part_y + 2
while sheet.cell(row=end_table, column=part_x).value: # and end_table < 1000:  # узнаем высоту таблицы
    end_table += 1
print(part_x, part_y)
print('Высота спецификации:' + str(end_table))
cleanDB()
print('Очистил базу данных')
insertDB(sheet)
print('Данные вставлены')

# db.session.query(Product.part, Service.sku, Service.serv_gpl).join(Service)\
#     .filter(Service.serv_lev == 'SNT', Product.part == '01FT600X').all()
