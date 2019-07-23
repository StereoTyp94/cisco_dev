import openpyxl, os
from app import app, db
from app.models import Product, Service

path = os.chdir('../price')
print(os.getcwd())
print(os.listdir())
wb = openpyxl.load_workbook(os.getcwd()+'\Report2.xlsx')
sheet = wb.get_active_sheet()
print(sheet)
for part_y in range(1, 20): #ищем начало таблицы и номера столбцов
    if 'Major' in str(sheet.cell(row=part_y, column=1).value):
        for part_x in range(1,15):
            if 'Product SKU' == str(sheet.cell(row=part_y, column=part_x).value):
                break
        for serv_lev in range(1,15):
            if 'Service Level' == str(sheet.cell(row=part_y, column=serv_lev).value):
                break
        for sku in range(1,15):
            if 'Service SKU' == str(sheet.cell(row=part_y, column=sku).value):
                gpl = sku + 1
                serv_gpl = gpl + 1
                break
        break
if part_y == 20:
    print('Не смог найти начало таблицы')
print(part_x,part_y)
end_table = part_y + 2
while sheet.cell(row=end_table, column=part_x).value and end_table < 30: #узнаем высоту таблицы
    end_table += 1
print(end_table)

part = ''
for line in range(part_y+2, end_table):
    if sheet.cell(row=line, column=part_x).value != part:
        part = sheet.cell(row=line, column=part_x).value
        if Product.query.filter_by(part=part).first() is None:
            product = Product(part=part, gpl=float(sheet.cell(row=line, column=gpl).value))
            db.session.add(product)
            db.session.commit()
        else:
            part_cur = Product.query.filter_by(part=part).first()
            part_cur.gpl = float(sheet.cell(row=line, column=gpl).value)
            db.session.commit()
    sku_name = sheet.cell(row=line, column=sku).value
    if Service.query.filter_by(sku=sku_name).first() is None:
        service = Service(serv_lev=sheet.cell(row=line, column=serv_lev).value, sku=sku_name,\
                          serv_gpl=float(sheet.cell(row=line, column=serv_gpl).value), equipment=Product.query.filter_by(part=part).first())
        db.session.add(service)
        db.session.commit()
    else:
        sku_cur = Service.query.filter_by(sku=sku_name).first()
        sku_cur.serv_gpl = float(sheet.cell(row=line, column=serv_gpl).value)
        db.session.commit()

# db.session.query(Product.part, Service.sku, Service.serv_gpl).join(Service)\
#     .filter(Service.serv_lev == 'SNT' and Product.part == '01FT600X').all()
